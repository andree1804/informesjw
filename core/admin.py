from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from .models import Report, Person, Group, Privilege, PrivilegePermanent
from .forms import GroupSelectForm

from .models import PersonVirtual
from django import forms
from django.db.models import Sum
from collections import OrderedDict
from .forms import GroupSelectForm
from django.urls import reverse
from datetime import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone
from django.utils.formats import date_format
from django.utils.timezone import localtime
from django.utils.safestring import mark_safe
from collections import defaultdict

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import base64
import calendar
import locale

from django.shortcuts import get_object_or_404
from pdfrw import PdfReader, PdfWriter, PdfDict, PdfObject, PdfName
import os
import tempfile
import zipfile
from django.db.models import Q


from django.conf import settings
#import requests

from .forms import PersonForm

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from django.db import connection


def export_to_xls(modeladmin, request, queryset):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personas"
    
    # Definir campos de contactos (hasta 3 contactos)
    contact_fields = [
        ('contact_apoderado_1', 'contact_telefono_1'),
        ('contact_apoderado_2', 'contact_telefono_2'),
        ('contact_apoderado_3', 'contact_telefono_3'),
    ]
    
    # Crear encabezados incluyendo los campos de contactos
    headers = [
        'ID', 'Nombres', 'Grupo', 'Privilegio', 
        'Privilegios Permanentes', 'Fecha Nacimiento', 
        'Fecha Bautismo', 'Género', 'Esperanza', 
        'Teléfono', 'Dirección'
    ]
    
    # Agregar encabezados de contactos
    for apoderado_field, telefono_field in contact_fields:
        headers.extend([f'Apoderado {apoderado_field[-1]}', f'Teléfono {telefono_field[-1]}'])
    
    # Aplicar formato a los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Llenar datos con el queryset seleccionado
    for row_num, person in enumerate(queryset, 2):
        # Obtener privilegios permanentes como string
        privileges_permanent = ", ".join([str(pp) for pp in person.privileges_permanent.all()])
        
        # Convertir campos booleanos a texto legible
        gender = 'Hombre' if person.gender else 'Mujer'
        hope = 'Ungido' if person.hope else 'Muchedumbre'
        
        # Datos básicos
        data_row = [
            person.id,
            person.names,
            str(person.group),
            str(person.privilege),
            privileges_permanent,
            person.birth,
            person.baptism,
            gender,
            hope,
            person.phone,
            person.address
        ]
        
        # Llenar datos básicos
        for col_num, value in enumerate(data_row, 1):
            ws.cell(row=row_num, column=col_num, value=value)
        
        # Llenar datos de contactos
        contact_col_start = len(data_row) + 1
        
        for contact_index, (apoderado_field, telefono_field) in enumerate(contact_fields):
            if contact_index < len(person.contacts):
                contact = person.contacts[contact_index]
                # Apoderado del contacto
                ws.cell(row=row_num, column=contact_col_start + (contact_index * 2), 
                       value=contact.get('apoderado', ''))
                # Teléfono del contacto
                ws.cell(row=row_num, column=contact_col_start + (contact_index * 2) + 1, 
                       value=contact.get('telefono', ''))
            else:
                # Campos vacíos si no hay contacto
                ws.cell(row=row_num, column=contact_col_start + (contact_index * 2), value='')
                ws.cell(row=row_num, column=contact_col_start + (contact_index * 2) + 1, value='')
    
    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="personas_export.xlsx"'
    
    # Guardar el libro de trabajo en la respuesta
    wb.save(response)
    
    return response

export_to_xls.short_description = "Exportar seleccionados a XLS"


class PersonAdmin(admin.ModelAdmin):
    form = PersonForm
    list_display = ('id', 'names', 'group', 'privilege', 'baptism')

    actions = [export_to_xls]

    def get_fieldsets(self, request, obj=None):
        fieldsets = [
            (None, {
                'fields': [
                    'names',
                    'group',
                    'privilege',
                    'privileges_permanent',
                    'birth',
                    'baptism',
                    'gender',
                    'hope',
                    'phone',
                    'address',
                ]
            })
        ]
        
        # Add contacts section
        contact_fields = [
            ('contact_name_0', 'contact_phone_0'),
            ('contact_name_1', 'contact_phone_1'),
            ('contact_name_2', 'contact_phone_2'),
        ]
        
        fieldsets.append(('Contactos de emergencia', {
            'fields': contact_fields,
            'classes': ('collapse',)  # Optional: makes this section collapsible
        }))
        
        return fieldsets


class ReportAdmin(admin.ModelAdmin):
    change_list_template = "admin/reportes_por_grupo.html"

    def changelist_view(self, request, extra_context=None):
        now = datetime.now()
        if now.month == 1:  # Si es enero
            previous_year = now.year - 1
        else:
            previous_year = now.year

        previous_month = now.month - 1
        persons = []
        group = []
        group_id = request.GET.get('group')
        selected_month = request.GET.get('month', previous_month)
        selected_year = request.GET.get('year', previous_year)

        form = GroupSelectForm({
            'month': selected_month,
            'year': selected_year,
            **request.GET
        })

        if request.user.username != 'admin':
            group_default = Group.objects.filter(name=request.user.username.capitalize()).first()
            groups = Group.objects.all()
        else:
            groups = Group.objects.all()
        if group_id and selected_month and selected_year:    
            # Obtener los reportes ya existentes
            reports = Report.objects.filter(
                group_id=group_id,
                #person__in=persons,
                month=selected_month,
                year=selected_year
            ).select_related('privilege')

            if reports:
                persons_list = []
                for r in reports:
                    person = Person.objects.filter(id=r.person_id)\
                        .select_related('group', 'privilege')\
                        .order_by('names').first()  # Usar first() ya que id es único
                    if person:
                        persons_list.append(person)
                persons = persons_list  # Ahora tienes todas las personas
            else:
                persons = list(Person.objects.filter(group_id=group_id).order_by('names'))

            report_map = {r.person_id: r for r in reports}
            
            for person in persons:
                person.existing_report = report_map.get(person.id)
                if person.existing_report:
                    person.privilege = person.existing_report.privilege  # <--- esta línea es CLAVE


        if request.method == "POST":
            group_id = request.POST.get('group')
            selected_month = request.POST.get('month')
            selected_year = request.POST.get('year')
            total = int(request.POST.get('total', 0))

            try:
                group = Group.objects.get(pk=group_id)
            except Group.DoesNotExist:
                messages.error(request, "Grupo no encontrado.")
                return redirect(request.path)

            for i in range(total):
                person_id = request.POST.get(f'person_{i}')
                if not person_id:
                    continue

                try:
                    person = Person.objects.get(pk=person_id)
                    privilege_id = request.POST.get(f'privilege_{i}')
                    privilege = Privilege.objects.get(pk=int(privilege_id)) if privilege_id else None

                    courses_str = request.POST.get(f'courses_{i}', '0')
                    hours_str = request.POST.get(f'hours_{i}', '0')
                    courses = int(courses_str) if courses_str.strip().isdigit() else 0
                    hours = int(hours_str) if hours_str.strip().isdigit() else 0
                    participated = f'participated_{i}' in request.POST
                    note = request.POST.get(f'note_{i}', '')

                    report, created = Report.objects.get_or_create(
                        person=person,
                        month=selected_month,
                        year=selected_year,
                        defaults={
                            'group': group,
                            'privilege': privilege,
                            'courses': courses,
                            'hours': hours,
                            'participated': participated,
                            'note': note
                        }
                    )

                    if not created:
                        report.group = group
                        report.privilege = privilege
                        report.courses = courses
                        report.hours = hours
                        report.participated = participated
                        report.note = note
                        report.save()
                except Exception as e:
                    messages.error(request, f"Error al guardar para ID {person_id}: {str(e)}")

            messages.success(request, "Datos guardados correctamente.")
            return redirect(f"{request.path}?group={group_id}&month={selected_month}&year={selected_year}")

        if request.user.username != 'admin':
            group_default = Group.objects.filter(name=request.user.username.capitalize()).first()
            groups = Group.objects.all()

            form.fields['group'].choices = [(group.id, group.name) for group in groups]
            
            if group_default:
                form.fields['group'].initial = group_default.id
        else:
            groups = Group.objects.exclude(name='-')
            form.fields['group'].choices = [(group.id, group.name) for group in groups]

        if persons:
            group_name = persons[0].group.name
        else:
            group_name = ''

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'persons': persons,
            'group_id': group_id,
            'mes_actual': selected_month,
            'año_actual': selected_year,
            'privilegios': Privilege.objects.all().order_by('id'),
            'group': group_id,
            'group_name': group_name,
            'month': selected_month,
            'year': selected_year,
            'opts': self.model._meta,  # Necesario para el admin
            'title': 'Informes de servicio',
        }

        if extra_context:
            context.update(extra_context)

        return render(request, self.change_list_template, context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'consolidado/',
                self.admin_site.admin_view(ConsolidatedAdmin(Report, self.admin_site).changelist_view),
                name='report_consolidated'
            ),
            path(
                'generar-pdf/',
                self.admin_site.admin_view(self.generar_pdf),
                name='report_generar_pdf',
            ),
            path(
                'generar-pdf-publicadores/',
                self.admin_site.admin_view(self.generar_pdf_publicadores),
                name='report_generar_pdf_publicadores',
            ),
        ]
        return custom_urls + urls

    def fill_pdf(self, person, reports, template_path, output_path, year_service):
        template = PdfReader(template_path)

        # Activar renderizado visual de campos
        if template.Root.AcroForm:
            template.Root.AcroForm.update(PdfDict(NeedAppearances=PdfObject('true')))

        for page in template.pages:
            annotations = page['/Annots']
            if not annotations:
                continue

            for annotation in annotations:
                field_raw = annotation.get('/T')
                if not field_raw:
                    continue

                # Decodificar UTF-16 (quitar þÿ)
                field = field_raw.to_unicode().replace('þÿ', '').strip()

                if person.birth:
                    birth = person.birth.strftime('%d/%m/%Y')
                
                if person.baptism:
                    baptism = person.baptism.strftime('%d/%m/%Y')
                # Campos individuales
                if field == '900_1_Text_SanSerif':
                    annotation.update(PdfDict(V=person.names))
                if field == '900_2_Text_SanSerif' and person.birth:
                    annotation.update(PdfDict(V=str(birth)))
                if field == '900_3_CheckBox' and person.gender == True:
                    annotation.update(PdfDict(AS=PdfName('Yes')))
                if field == '900_4_CheckBox' and person.gender == False:
                    annotation.update(PdfDict(AS=PdfName('Yes')))
                if field == '900_5_Text_SanSerif' and person.baptism:
                    annotation.update(PdfDict(V=str(baptism)))
                if field == '900_6_CheckBox' and person.hope == False:
                    annotation.update(PdfDict(AS=PdfName('Yes')))
                if field == '900_7_CheckBox' and person.hope == True:
                    annotation.update(PdfDict(AS=PdfName('Yes')))

                for priv in person.privileges_permanent.all():
                    if field == '900_8_CheckBox' and priv.name == 'Anciano':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    if field == '900_9_CheckBox' and priv.name == 'Siervo Ministerial':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    if field == '900_10_CheckBox' and priv.name == 'Precursor Regular':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    if field == '900_11_CheckBox' and priv.name == 'Precursor Especial':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    if field == '900_12_CheckBox' and priv.name == 'Misionero':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                if field == '900_13_Text_C_SanSerif':
                    annotation.update(PdfDict(V=year_service))
                    #annotation.update(PdfDict(V=f"{(reports[0].year)-1}-{int(reports[0].year)}"))

                # Meses dinámicos
                for report in reports:
                    #m = int(report.month)
                    total_hours = sum(int(r.hours or 0) for r in reports)
                    if int(report.month) == 9:
                        m = 1
                    elif int(report.month) == 10:
                        m = 2
                    elif int(report.month) == 11:
                        m = 3
                    elif int(report.month) == 12:
                        m = 4
                    elif int(report.month) == 8:
                        m = 12
                    elif int(report.month) == 7:
                        m = 11
                    elif int(report.month) == 6:
                        m = 10
                    elif int(report.month) == 5:
                        m = 9
                    elif int(report.month) == 4:
                        m = 8
                    elif int(report.month) == 3:
                        m = 7
                    elif int(report.month) == 2:
                        m = 6
                    elif int(report.month) == 1:
                        m = 5

                    if field == f'901_{20 + m - 1}_CheckBox' and report.participated:
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    elif field == f'902_{20 + m - 1}_Text_C_SanSerif' and report.courses > 0:
                        annotation.update(PdfDict(V=str(report.courses)))
                    elif field == f'903_{20 + m - 1}_CheckBox' and report.privilege.name == 'Auxiliar':
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    elif field == f'904_{20 + m - 1}_S21_Value' and report.hours > 0:
                        annotation.update(PdfDict(V=str(report.hours)))
                    elif field == f'905_{20 + m - 1}_Text_SanSerif' and report.note:
                        annotation.update(PdfDict(V=str(report.note)))

                if field == '904_32_S21_Value' and total_hours > 0:
                    annotation.update(PdfDict(V=str(total_hours)))  # o el valor dinámico que desees

        PdfWriter().write(output_path, template)

    def generar_pdf(self, request):
        person_id = request.GET.get('person')
        group_id = request.GET.get('group')
        month = request.GET.get('month')
        year = request.GET.get('year')

        # Determinar el año teocrático al que pertenece el mes seleccionado
        year = int(year)
        if int(month) >= 9:
            start_year = year
            end_year = int(year) + 1
            year_service = f"{year}-{int(year + 1)}"
        else:
            start_year = int(year) - 1
            end_year = year
            year_service = f"{int(year - 1)}-{year}"

        template_path = os.path.join(settings.BASE_DIR, 'publisher_cards/pdf_base.pdf')  # <-- actualiza la ruta

        if person_id:
            person = get_object_or_404(Person, pk=person_id)
            # Filtrado de reportes en el año teocrático
            reports = Report.objects.filter(
                person=person
            ).filter(
                Q(year=start_year, month__in=['9', '10', '11', '12']) |  # Septiembre–Diciembre del año inicial
                Q(year=end_year, month__in=['1', '2', '3', '4', '5', '6', '7', '8'])  # Enero–Agosto del siguiente año
            ).order_by('year', 'month')

            output_path = os.path.join(tempfile.gettempdir(), f"tarjeta_{person.id}.pdf")
            self.fill_pdf(person, list(reports), template_path, output_path)

            with open(output_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename=tarjeta_{person.names}.pdf'
                return response

        elif group_id:
            group = get_object_or_404(Group, pk=group_id)
            persons = Person.objects.filter(group=group).order_by('names')

            temp_dir = tempfile.mkdtemp()
            zip_filename = os.path.join(temp_dir, f"tarjetas_grupo_{group.id}.zip")

            with zipfile.ZipFile(zip_filename, 'w') as zipf:
                for person in persons:
                    # Filtrado de reportes en el año teocrático
                    reports = Report.objects.filter(
                        person=person
                    ).filter(
                        Q(year=start_year, month__in=['9', '10', '11', '12']) |  # Septiembre–Diciembre del año inicial
                        Q(year=end_year, month__in=['1', '2', '3', '4', '5', '6', '7', '8'])  # Enero–Agosto del siguiente año
                    ).order_by('year', 'month')

                    if reports.exists():
                        if person.privilege.name == 'Publicador':
                            priv_name = ''
                        else:
                            priv_name = f"-{person.privilege.name}"
                        pdf_path = os.path.join(temp_dir, f"{person.names}{priv_name}.pdf")
                        self.fill_pdf(person, list(reports), template_path, pdf_path, year_service)
                        zipf.write(pdf_path, arcname=f"{person.names}{priv_name}.pdf")

            with open(zip_filename, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename=tarjetas_{group.name}.zip'
                return response

        else:
            return HttpResponse("Debes enviar al menos 'person' o 'group' en la URL.", status=400)

    def generar_pdf_publicadores(self, request):
        month = request.GET.get('month')
        year = request.GET.get('year')
        privilege = request.GET.get('privilege')

        # Determinar el año teocrático al que pertenece el mes seleccionado
        year = int(year)
        if int(month) >= 9:
            start_year = year
            end_year = int(year) + 1
            year_service = f"{year}-{int(year + 1)}"
        else:
            start_year = int(year) - 1
            end_year = year
            year_service = f"{int(year - 1)}-{year}"

        if not month or not year:
            return HttpResponse("Debes enviar al menos 'month' y 'year' en la URL.", status=400)

        if privilege == 'Publicador':
            template_path = os.path.join(settings.BASE_DIR, 'publisher_cards/pdf_base_publicadores.pdf')
            name_pdf = '1-Publicadores'
        elif privilege == 'Auxiliar':
            template_path = os.path.join(settings.BASE_DIR, 'publisher_cards/pdf_base_auxiliares.pdf')
            name_pdf = '2-Auxiliares'
        elif privilege == 'PR':
            template_path = os.path.join(settings.BASE_DIR, 'publisher_cards/pdf_base_regulares.pdf')
            name_pdf = '3-Regulares'
        
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, "tarjeta_publicadores.pdf")

        reports = Report.objects.filter(
            privilege__name=privilege,
            participated=True
        ).filter(
            Q(year=start_year, month__in=['9', '10', '11', '12']) |  # Septiembre–Diciembre del año inicial
            Q(year=end_year, month__in=['1', '2', '3', '4', '5', '6', '7', '8'])  # Enero–Agosto del siguiente año
        ).order_by('year', 'month')
        #reports = Report.objects.filter(year__in=years, privilege__name=privilege).order_by('month')

        if not reports.exists():
            return HttpResponse("No se encontraron reportes para esos meses.")

        self.fill_pdf_publicadores(list(reports), template_path, output_path, privilege, year_service)

        with open(output_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{name_pdf}_{year_service}.pdf"'
            return response

    def fill_pdf_publicadores(self, reports, template_path, output_path, privilege, year_service):
        template = PdfReader(template_path)

        if template.Root.AcroForm:
            template.Root.AcroForm.update(PdfDict(NeedAppearances=PdfObject('true')))

        # Agrupar por mes y acumular valores
        acumulado_por_mes = defaultdict(lambda: {
            'participated': 0,
            'courses': 0,
            'auxiliares': 0,
            'hours': 0,
            'notes': 0
        })

        for r in reports:
            acumulado = acumulado_por_mes[int(r.month)]
            acumulado['participated'] += 1 if r.participated else 0
            acumulado['courses'] += r.courses or 0
            acumulado['hours'] += r.hours or 0
            acumulado['auxiliares'] += 1 if r.privilege and r.privilege.name == 'Auxiliar' else 0
            acumulado['notes'] += r.participated or 0

        # Total general de horas
        total_hours = sum(d['hours'] for d in acumulado_por_mes.values())

        # Mapear mes calendario a columna PDF (1 → 5, ..., 9 → 1)
        def mes_a_columna_pdf(mes, privilege):
            return {9: 1, 10: 2, 11: 3, 12: 4, 1: 5, 2: 6, 3: 7, 4: 8, 5: 9, 6: 10, 7: 11, 8: 12}.get(mes)

        for page in template.pages:
            annotations = page['/Annots']
            if not annotations:
                continue

            for annotation in annotations:
                field_raw = annotation.get('/T')
                if not field_raw:
                    continue

                field = field_raw.to_unicode().replace('þÿ', '').strip()

                if field == '900_10_CheckBox' and privilege == 'PR':
                    annotation.update(PdfDict(AS=PdfName('Yes')))

                if field == '900_13_Text_C_SanSerif':
                    annotation.update(PdfDict(V=year_service))

                # Revisar todos los meses acumulados
                for mes, datos in acumulado_por_mes.items():
                    columna = mes_a_columna_pdf(mes, privilege)
                    if not columna:
                        continue

                    if field == f'901_{20 + columna - 1}_CheckBox' and datos['participated'] > 0:
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    elif field == f'902_{20 + columna - 1}_Text_C_SanSerif' and datos['courses'] > 0:
                        annotation.update(PdfDict(V=str(datos['courses'])))
                    elif field == f'903_{20 + columna - 1}_CheckBox' and datos['auxiliares'] > 0:
                        annotation.update(PdfDict(AS=PdfName('Yes')))
                    elif field == f'904_{20 + columna - 1}_S21_Value' and datos['hours'] > 0:
                        annotation.update(PdfDict(V=str(datos['hours'])))
                    elif field == f'905_{20 + columna - 1}_Text_SanSerif' and datos['notes']:
                        annotation.update(PdfDict(V=str(datos['notes'])))
                        #annotation.update(PdfDict(V=" | ".join(datos['notes'])))

                if field == '904_32_S21_Value':
                    annotation.update(PdfDict(V=str(total_hours)))

        PdfWriter().write(output_path, template)


@admin.register(PersonVirtual)
class MeetingAttendanceAdmin(admin.ModelAdmin):
    list_display = ('formatted_date', 'in_person', 'virtual', 'total_attendance')
    list_filter = ('meeting_date',)
    date_hierarchy = 'meeting_date'
    change_list_template = 'admin/meeting_attendance_change_list.html'
    
    def changelist_view(self, request, extra_context=None):
        # Obtener el mes y año actual
        now = datetime.now()
        current_month = request.GET.get('month_', str(now.month))
        current_year = request.GET.get('year_', str(now.year))
        
        # Si no hay parámetros GET, redirigir con los valores por defecto
        if not request.GET.get('month_') and not request.GET.get('year_'):
            base_url = reverse('admin:core_personvirtual_changelist')
            return redirect(f'{base_url}?year_={current_year}&month_={current_month}')
        
        form = GroupSelectForm({
            'month': current_month,
            'year': current_year,
            **request.GET
        })
        
        # Resto de tu lógica existente...
        registros = PersonVirtual.objects.filter(
            meeting_date__month=current_month,
            meeting_date__year=current_year
        ).order_by('meeting_date')

        meses = OrderedDict()
        for reg in registros:
            fecha = reg.meeting_date
            mes = fecha.strftime("%B") if fecha else "Sin fecha"
            if mes not in meses:
                meses[mes] = {'registros': [], 'subtotal': 0}
            meses[mes]['registros'].append(reg)
            total = (reg.in_person or 0) + (reg.virtual or 0)
            meses[mes]['subtotal'] += total

        totales = PersonVirtual.objects.aggregate(
            total_in_person=Sum('in_person'),
            total_virtual=Sum('virtual')
        )
        total_general = (totales['total_in_person'] or 0) + (totales['total_virtual'] or 0)

        context = {
            **self.admin_site.each_context(request),
            'meses': meses,
            'total_in_person': totales['total_in_person'] or 0,
            'total_virtual': totales['total_virtual'] or 0,
            'total_general': total_general,
            'form': form,
            'opts': self.model._meta,  # Necesario para el admin
            'title': 'Asistencia a las reuniones'
        }

        if extra_context:
            context.update(extra_context)

        return render(request, self.change_list_template, context)

    def formatted_date(self, obj):
        return obj.meeting_date.strftime("%A-%d") if obj.meeting_date else "-"
    formatted_date.short_description = 'Fecha'
    
    def total_attendance(self, obj):
        return (obj.in_person or 0) + (obj.virtual or 0)
    total_attendance.short_description = 'Total'

    def save_model(self, request, obj, form, change):
        if obj.in_person is None:
            obj.in_person = 0
        if obj.virtual is None:
            obj.virtual = 0
        super().save_model(request, obj, form, change)


class ConsolidatedAdmin(admin.ModelAdmin):
    list_display = ('person', 'group', 'privilege', 'month', 'year')
    list_filter = ('month', 'year', 'group', 'privilege')
    search_fields = ('person__name',)
    
    change_list_template = 'admin/reports_consolidated.html'
    
    def changelist_view(self, request, extra_context=None):
        if 'generate_report' in request.GET:
            month = request.GET.get('month')
            year = request.GET.get('year')
            
            # Obtener datos consolidados
            reports = Report.objects.filter(month=month, year=year)
            person_virtual = PersonVirtual.objects.filter(meeting_date__month=month, meeting_date__year=year)

            # Filtrar solo sábados y domingos
            fines_de_semana = [
                pv for pv in person_virtual
                if pv.meeting_date.weekday() in (5, 6)
            ]

            # Calcular asistencia total en fines de semana
            total_asistencia = sum(
                (pv.in_person or 0) + (pv.virtual or 0)
                for pv in fines_de_semana
            )

            # Calcular asistencia promedio
            cantidad_reuniones = len(fines_de_semana)
            asistencia_promedio = (
                total_asistencia / cantidad_reuniones
                if cantidad_reuniones > 0 else 0
            )
            
            # Calcular estadísticas
            stats = {
                'publicadores': {
                    'count': reports.filter(privilege__name='Publicador', participated=True).count(),
                    'courses': reports.filter(privilege__name='Publicador').aggregate(Sum('courses'))['courses__sum'] or 0,
                },
                'auxiliares': {
                    'count': reports.filter(privilege__name='Auxiliar', participated=True).count(),
                    'hours': reports.filter(privilege__name='Auxiliar').aggregate(Sum('hours'))['hours__sum'] or 0,
                    'courses': reports.filter(privilege__name='Auxiliar').aggregate(Sum('courses'))['courses__sum'] or 0,
                },
                'regulares': {
                    'count': reports.filter(privilege__name='PR', participated=True).count(),
                    'hours': reports.filter(privilege__name='PR').aggregate(Sum('hours'))['hours__sum'] or 0,
                    'courses': reports.filter(privilege__name='PR').aggregate(Sum('courses'))['courses__sum'] or 0,
                },
                'total_activos': reports.filter(participated=True).count(),
                'asistencia_promedio': round(asistencia_promedio, 1),
                'month': month,
                'year': year,
                'updated_by': request.user.get_full_name(),
                'updated_date': date_format(localtime(), format='j \d\e F \d\e Y', use_l10n=True)
            }
            
            if request.GET.get('format') == 'pdf':
                return self.generate_pdf_report(stats)
            else:
                extra_context = {'stats': stats}
        
        return super().changelist_view(request, extra_context=extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'reportes_por_grupo/',
                self.admin_site.admin_view(ReportAdmin(Report, self.admin_site).changelist_view),
                name='reportes_por_grupo'
            ),
        ]
        return custom_urls + urls
    
    def generate_pdf_report(self, stats):
        #locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        MESES_ES = [
            "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
        ]
        month_es = MESES_ES[int(stats['month']) - 1]

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="informe_{stats["month"]}_{stats["year"]}.pdf"'
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        
        # Configuración inicial
        p.setTitle(f"Informe Consolidado - {month_es} {stats['year']}")
        
        # Encabezado
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, "Informes de la congregación - Pedregal")
        p.setFont("Helvetica", 12)
        p.drawString(100, 750, "PREDICACIÓN Y ASISTENCIA A LAS REUNIONES")
        
        # Fecha y actualización
        p.drawString(100, 730, f"{month_es.capitalize()} de {stats['year']}")
        p.drawString(100, 715, f"Actualizado por: {stats['updated_by']}")
        p.drawString(100, 700, f"Actualizado el: {stats['updated_date']}")
        
        # Totales
        p.drawString(100, 685, "Todos los publicadores activos")
        p.drawString(400, 685, str(stats['total_activos']))
        # Asistencia promedio solo fin de semana
        p.drawString(100, 670, "Promedio asistencia a reuniones de fin de semana")
        p.drawString(400, 670, str(stats['asistencia_promedio']))

        # Publicadores
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 650, "Publicadores")
        p.setFont("Helvetica", 12)
        p.drawString(120, 635, "Cantidad de informes")
        p.drawString(400, 635, str(stats['publicadores']['count']))
        p.drawString(120, 620, "Cursos bíblicos")
        p.drawString(400, 620, str(stats['publicadores']['courses']))
        
        # Precursores auxiliares
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 590, "Precursores auxiliares")
        p.setFont("Helvetica", 12)
        p.drawString(120, 575, "Cantidad de informes")
        p.drawString(400, 575, str(stats['auxiliares']['count']))
        p.drawString(120, 560, "Horas")
        p.drawString(400, 560, str(stats['auxiliares']['hours']))
        p.drawString(120, 545, "Cursos bíblicos")
        p.drawString(400, 545, str(stats['auxiliares']['courses']))
        
        # Precursores regulares
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 515, "Precursores regulares")
        p.setFont("Helvetica", 12)
        p.drawString(120, 500, "Cantidad de informes")
        p.drawString(400, 500, str(stats['regulares']['count']))
        p.drawString(120, 485, "Horas")
        p.drawString(400, 485, str(stats['regulares']['hours']))
        p.drawString(120, 470, "Cursos bíblicos")
        p.drawString(400, 470, str(stats['regulares']['courses']))
        
        p.showPage()
        p.save()
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response


admin.site.register(Report, ReportAdmin)
admin.site.register(Group)
admin.site.register(Privilege)
admin.site.register(PrivilegePermanent)
admin.site.register(Person, PersonAdmin)



